#!/usr/bin/env python3
"""
export_excel.py  —  Export CM Hammar product catalogue to a multi-sheet Excel file.

Each sheet corresponds to a top-level category (H20, Remote Release Systems, etc.).
Columns are all product fields + flattened structured data from data sheets and
approval certificates. Each row is one product.

Usage:
    python export_excel.py
    python export_excel.py --output my_catalogue.xlsx
    python export_excel.py --depth 2   # one sheet per sub-category instead
"""

import argparse
import json
import os
import re
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, GradientFill, PatternFill, Side
)
from openpyxl.utils import get_column_letter

OUT_DIR   = os.path.join(os.path.dirname(__file__), "output")
HIER_PATH = os.path.join(OUT_DIR, "hierarchy.json")
OUT_XLSX  = os.path.join(OUT_DIR, "cm_hammar_catalogue.xlsx")

# ── Colours ──────────────────────────────────────────────────────────────────
NAVY       = "0C1F33"
TEAL       = "1980A0"
TEAL_LIGHT = "DEF0F6"
GREEN      = "1A7A4A"   # section divider tabs for detailed sheets
HEADER_FG  = "FFFFFF"
ALT_ROW    = "F2F7FA"
BORDER_CLR = "D0DDE8"
WARN_BG    = "FFF3CD"
WARN_FG    = "7B5A00"

# ── Styles ───────────────────────────────────────────────────────────────────
def hdr_font():  return Font(bold=True, color=HEADER_FG, size=10, name="Calibri")
def hdr_fill():  return PatternFill("solid", fgColor=NAVY)
def sub_fill():  return PatternFill("solid", fgColor=TEAL)
def alt_fill():  return PatternFill("solid", fgColor=ALT_ROW)
def thin_border():
    s = Side(border_style="thin", color=BORDER_CLR)
    return Border(left=s, right=s, top=s, bottom=s)
def cell_font(): return Font(size=9, name="Calibri")
def wrap_align(): return Alignment(wrap_text=True, vertical="top")
def center_align(): return Alignment(horizontal="center", vertical="top")

# ── Value flattening ──────────────────────────────────────────────────────────

def flat_str(val, sep=" | ") -> str:
    """Convert any structured value to a readable string."""
    if val is None:
        return ""
    if isinstance(val, str):
        return val.strip()
    if isinstance(val, list):
        parts = []
        for item in val:
            if isinstance(item, dict):
                parts.append(flat_str(item, sep))
            else:
                parts.append(str(item))
        return sep.join(p for p in parts if p)
    if isinstance(val, dict):
        # {label, value} variant object
        if "label" in val and "value" in val:
            return f"{val['label']}: {val['value']}"
        # {variants, notes} object
        if "variants" in val:
            parts = [flat_str(v) for v in val["variants"]]
            if val.get("notes"):
                parts.append(flat_str(val["notes"]))
            return sep.join(p for p in parts if p)
        # Generic dict: key: value pairs
        parts = []
        for k, v in val.items():
            if k == "notes":
                notes = v if isinstance(v, list) else [v]
                parts.extend(str(n) for n in notes)
            else:
                parts.append(f"{k}: {flat_str(v)}")
        return sep.join(p for p in parts if p)
    return str(val)


def flatten_section(heading: str, val) -> dict:
    """
    Flatten a structured section into {column_name: cell_value} pairs.
    For variant-heavy sections (HYDROSTATIC etc.) each label becomes its own column.
    For other sections a single column per heading is used.
    """
    prefix = heading.title().replace(" ", " ")

    if val is None:
        return {prefix: ""}

    # Flat array of {label, value} → one column per label
    if (isinstance(val, list) and val
            and isinstance(val[0], dict)
            and "label" in val[0]):
        return {f"{prefix} – {v.get('label','')}" : v.get("value","") for v in val}

    # {variants, notes} → expand variants + notes
    if isinstance(val, dict) and "variants" in val:
        result = {}
        for v in val.get("variants", []):
            result[f"{prefix} – {v.get('label','')}"] = v.get("value", "")
        if val.get("notes"):
            result[f"{prefix} – Notes"] = flat_str(val["notes"], " ")
        return result

    # Nested dict {row_label: {component: value}} — SPECIFICATIONS table
    if isinstance(val, dict) and val and isinstance(next(iter(val.values())), dict):
        result = {}
        for row_label, components in val.items():
            if isinstance(components, dict):
                for component, cell_val in components.items():
                    col = f"{prefix}: {row_label} ({component})"
                    result[col] = flat_str(cell_val)
            else:
                result[f"{prefix}: {row_label}"] = flat_str(components)
        return result

    # Plain dict (key-value + optional notes) → flatten
    if isinstance(val, dict):
        result = {}
        for k, v in val.items():
            if k == "notes":
                result[f"{prefix} – Notes"] = flat_str(v, " ")
            else:
                result[f"{prefix} – {k}"] = flat_str(v)
        return result

    # Everything else → single column
    return {prefix: flat_str(val)}


def extract_product_row(product: dict) -> dict:
    """
    Build a flat dict of all columns for one product row.
    Order: core fields → product specs → data sheet → approvals.
    """
    row = OrderedDict()

    # ── Core fields ───────────────────────────────────────────────
    row["SKU"]        = product.get("sku", "")
    row["Name"]       = product.get("name", "")
    row["Subtitle"]   = product.get("subtitle", "")
    row["URL"]        = product.get("url", "")
    row["Image URL"]  = product.get("image", "")
    row["Categories"] = " | ".join(product.get("categories", []))
    row["Tags"]       = " | ".join(product.get("tags", []))

    # ── Product specifications (scraped from page) ─────────────────
    for k, v in (product.get("specifications") or {}).items():
        row[f"Spec: {k}"] = str(v) if v is not None else ""

    # ── Notes ─────────────────────────────────────────────────────
    if product.get("notes"):
        row["Notes"] = product["notes"]

    # ── Data sheet structured data ─────────────────────────────────
    ds = None
    for item in (product.get("downloads") or []):
        s = item.get("structured") or {}
        if isinstance(s, dict) and s.get("document_type") == "data_sheet":
            ds = s; break

    if ds:
        row["DS: HS Code"]           = ds.get("hs_code") or ""
        row["DS: Country of Origin"] = ds.get("country_of_origin") or ""
        for heading, val in (ds.get("sections") or {}).items():
            row.update(flatten_section(heading, val))
        if ds.get("warnings"):
            row["DS: Warnings"] = " | ".join(ds["warnings"])

    # ── Approval certificates ──────────────────────────────────────
    cert_num = 1
    for item in (product.get("approvals") or []) + (product.get("downloads") or []):
        s = item.get("structured") or {}
        if not isinstance(s, dict):
            continue
        if s.get("document_type") not in ("approval", "approval_certificate"):
            continue
        p = f"Cert {cert_num}: "
        row[p + "Number"]         = s.get("certificate_number") or ""
        row[p + "Issuing Body"]   = s.get("issuing_body") or ""
        row[p + "Type"]           = s.get("certificate_type") or ""
        row[p + "Valid Until"]    = s.get("valid_until") or ""
        row[p + "Product"]        = s.get("product_name") or ""
        row[p + "Designation"]    = s.get("product_designation") or ""
        row[p + "Approved Versions"] = " | ".join(s.get("approved_versions") or [])
        row[p + "Regulations"]    = " | ".join(s.get("regulations") or [])
        row[p + "Manufacturer"]   = s.get("manufacturer") or ""
        if s.get("notes"):
            row[p + "Notes"]      = str(s["notes"])
        cert_num += 1

    return row


# ── Hierarchy walking ─────────────────────────────────────────────────────────

def collect_sheets(root: dict, depth: int) -> list[tuple[str, list[dict]]]:
    """
    Walk the hierarchy and return a list of (sheet_title, [products]) pairs.
    depth=1 → top-level categories; depth=2 → sub-categories; etc.
    """
    sheets = []

    def walk(node: dict, current_depth: int, path_titles: list[str]):
        title    = node.get("title", "Category")
        products = node.get("products") or []
        children = node.get("children") or {}

        if current_depth == depth or not children:
            # Collect all products under this node (including from all children)
            all_prods = []
            seen = set()
            def gather(n):
                for p in (n.get("products") or []):
                    if p["sku"] not in seen:
                        seen.add(p["sku"])
                        all_prods.append(p)
                for child in (n.get("children") or {}).values():
                    gather(child)
            gather(node)
            if all_prods:
                sheets.append((title, all_prods))
        else:
            # Recurse into children
            for child in children.values():
                walk(child, current_depth + 1, path_titles + [title])

    for child in (root.get("children") or {}).values():
        walk(child, 1, [])

    return sheets


def safe_sheet_name(name: str) -> str:
    """Excel sheet names: max 31 chars, no [ ] : * ? / \\"""
    name = re.sub(r'[\[\]:*?/\\]', '-', name)
    return name[:31]


# ── Excel writing ─────────────────────────────────────────────────────────────

def write_sheet(ws, products: list[dict], sheet_title: str):
    """Write one sheet: header row then one row per product."""

    # Collect all rows first to discover full column set
    rows = [extract_product_row(p) for p in products]

    # Build master column list preserving insertion order
    all_cols: list[str] = []
    seen_cols: set[str] = set()
    for row in rows:
        for col in row:
            if col not in seen_cols:
                all_cols.append(col)
                seen_cols.add(col)

    # ── Header row ────────────────────────────────────────────────
    ws.append(all_cols)
    hrow = ws[1]
    for cell in hrow:
        cell.font      = hdr_font()
        cell.fill      = hdr_fill()
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border    = thin_border()

    # ── Data rows ─────────────────────────────────────────────────
    for i, row in enumerate(rows, start=2):
        values = [row.get(col, "") for col in all_cols]
        ws.append(values)
        fill = alt_fill() if i % 2 == 0 else None
        for j, cell in enumerate(ws[i]):
            cell.font      = cell_font()
            cell.alignment = wrap_align()
            cell.border    = thin_border()
            if fill:
                cell.fill = fill
            # Highlight warning columns in amber
            if all_cols[j].startswith("DS: Warnings") or all_cols[j].startswith("Cert") and "Notes" in all_cols[j]:
                if cell.value:
                    cell.fill = PatternFill("solid", fgColor=WARN_BG)
                    cell.font = Font(size=9, name="Calibri", color=WARN_FG)

    # ── Column widths ─────────────────────────────────────────────
    col_widths = {
        "SKU": 14, "Name": 28, "Subtitle": 28, "URL": 22, "Image URL": 22,
        "Categories": 22, "Tags": 18, "Notes": 28,
    }
    for col_idx, col_name in enumerate(all_cols, start=1):
        letter = get_column_letter(col_idx)
        if col_name in col_widths:
            ws.column_dimensions[letter].width = col_widths[col_name]
        elif col_name.startswith("DS:") or col_name.startswith("Spec:"):
            ws.column_dimensions[letter].width = 22
        elif "Hydrostatic" in col_name or "Approved" in col_name or "Label And" in col_name:
            ws.column_dimensions[letter].width = 30
        elif col_name.startswith("Cert"):
            ws.column_dimensions[letter].width = 20
        else:
            ws.column_dimensions[letter].width = 18

    # ── Freeze header + SKU column ────────────────────────────────
    ws.freeze_panes = "B2"

    # ── Sheet tab colour ──────────────────────────────────────────
    ws.sheet_properties.tabColor = TEAL


def write_overview(ws, family_sheets: list[tuple[str, list[dict]]], detail_sheets: list[tuple[str, list[dict]]] | None = None):
    """Write an overview / index sheet listing all sheets and product counts."""
    ws.title = "Overview"
    ws.sheet_properties.tabColor = NAVY

    headers = ["Sheet", "Level", "Products", "Description"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = hdr_font()
        cell.fill = hdr_fill()
        cell.alignment = center_align()
        cell.border = thin_border()

    def _row(ws, i, title, level, prods):
        ds_count   = sum(1 for p in prods for item in (p.get("downloads") or [])
                         if (item.get("structured") or {}).get("document_type") == "data_sheet")
        cert_count = sum(1 for p in prods for item in (p.get("approvals") or [])
                         if (item.get("structured") or {}).get("document_type") in ("approval", "approval_certificate"))
        desc = f"{ds_count} data sheets, {cert_count} certificates"
        ws.append([title, level, len(prods), desc])
        fill = PatternFill("solid", fgColor="E8F4E8") if level == "Sub-category" else (alt_fill() if i % 2 == 0 else None)
        for cell in ws[i]:
            cell.font      = cell_font()
            cell.alignment = wrap_align()
            cell.border    = thin_border()
            if fill:
                cell.fill = fill

    row_idx = 2
    for title, prods in family_sheets:
        _row(ws, row_idx, title, "Product Family", prods)
        row_idx += 1

    if detail_sheets:
        for title, prods in detail_sheets:
            _row(ws, row_idx, title, "Sub-category", prods)
            row_idx += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 35
    ws.freeze_panes = "A2"


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output", default=OUT_XLSX, help="Output .xlsx path")
    parser.add_argument("--depth",  type=int, default=1,
                        help="Hierarchy depth for sheet grouping (1=top-level, 2=sub-categories)")
    parser.add_argument("--combined", action="store_true",
                        help="Include both product family sheets AND sub-category sheets in one workbook")
    args = parser.parse_args()

    print(f"Loading {HIER_PATH} …")
    with open(HIER_PATH, encoding="utf-8") as f:
        hierarchy = json.load(f)

    family_sheets = collect_sheets(hierarchy, 1)
    detail_sheets = collect_sheets(hierarchy, 2) if args.combined else None

    sheets = family_sheets if not args.combined else family_sheets
    print(f"Found {len(family_sheets)} product family sheets:")
    for title, prods in family_sheets:
        print(f"  [{len(prods):3d} products] {title}")
    if detail_sheets:
        print(f"Found {len(detail_sheets)} sub-category sheets:")
        for title, prods in detail_sheets:
            print(f"  [{len(prods):3d} products]   > {title}")

    wb = Workbook()
    wb.remove(wb.active)

    # Overview
    ws_ov = wb.create_sheet("Overview")
    write_overview(ws_ov, family_sheets, detail_sheets)

    # Product family sheets (teal tabs)
    for title, prods in family_sheets:
        name = safe_sheet_name(title)
        ws = wb.create_sheet(name)
        ws.title = name
        ws.sheet_properties.tabColor = TEAL
        print(f"  Writing '{name}' ({len(prods)} products) …")
        write_sheet(ws, prods, title)

    # Detailed sub-category sheets (green tabs, indented names)
    if detail_sheets:
        seen_names: dict[str, int] = {}
        for title, prods in detail_sheets:
            base = safe_sheet_name(title)
            # Deduplicate sheet names that clash across families
            if base in seen_names:
                seen_names[base] += 1
                name = safe_sheet_name(f"{base} ({seen_names[base]})")
            else:
                seen_names[base] = 1
                name = base
            ws = wb.create_sheet(name)
            ws.title = name
            ws.sheet_properties.tabColor = GREEN
            print(f"  Writing '{name}' ({len(prods)} products) …")
            write_sheet(ws, prods, title)

    wb.save(args.output)
    print(f"\nSaved: {args.output}")
    size_kb = os.path.getsize(args.output) // 1024
    print(f"File size: {size_kb} KB")


if __name__ == "__main__":
    main()
